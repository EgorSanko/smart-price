# -*- coding: utf-8 -*-
"""Read Антиплагиат 2026.xlsx and find Санько Е.А."""
import io
import openpyxl

SRC = r"C:/Users/egor3/Downloads/Антиплагиат 2026.xlsx"
wb = openpyxl.load_workbook(SRC, data_only=True)

out = io.StringIO()
out.write(f"Sheets: {wb.sheetnames}\n\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    out.write(
        f"=== Sheet: {sheet_name} (rows={ws.max_row}, cols={ws.max_column}) ===\n"
    )
    for row_idx in range(1, ws.max_row + 1):
        row_values = []
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            row_values.append(str(val) if val is not None else "")
        if any(v.strip() for v in row_values):
            out.write(f'  R{row_idx}: {" | ".join(row_values)}\n')
    out.write("\n")

with open(
    r"C:/Users/egor3/Desktop/smart-price/_xlsx_dump.txt", "w", encoding="utf-8"
) as f:
    f.write(out.getvalue())
print("OK")
